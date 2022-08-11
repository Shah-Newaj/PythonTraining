import openpyxl

# to save same data
# File--> Workbook--> Sheets--> Rows--> Cells

# file = "C:\Selenium Practice\WriteData.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet1"]
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Welcome"
#
# workbook.save(file)

#to save Multiple data
file = "C:\Selenium Practice\WriteData.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Shah"
sheet.cell(1,3).value="Cse"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="Newaj"
sheet.cell(2,3).value="MSc"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="OTher"
sheet.cell(3,3).value="BSc"

workbook.save(file)